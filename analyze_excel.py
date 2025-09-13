#!/usr/bin/env python3
"""
Script to analyze the current Financial Dashboard Excel file structure
"""
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import sys
import os

def analyze_excel_file(file_path):
    """Analyze the structure of an Excel file"""
    print(f"Analyzing Excel file: {file_path}")
    print("=" * 60)
    
    try:
        # Load the workbook
        wb = load_workbook(file_path, read_only=True)
        
        print(f"Total sheets: {len(wb.sheetnames)}")
        print(f"Sheet names: {wb.sheetnames}")
        print()
        
        # Analyze each sheet
        for i, sheet_name in enumerate(wb.sheetnames, 1):
            print(f"{i}. Sheet: '{sheet_name}'")
            print("-" * 40)
            
            # Load sheet data
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                print(f"   Dimensions: {df.shape[0]} rows × {df.shape[1]} columns")
                
                if not df.empty:
                    print(f"   Column headers: {list(df.columns[:10])}{'...' if len(df.columns) > 10 else ''}")
                    
                    # Show first few rows of data (non-empty cells)
                    non_empty_data = df.dropna(how='all').head(3)
                    if not non_empty_data.empty:
                        print("   Sample data (first 3 non-empty rows):")
                        for idx, row in non_empty_data.iterrows():
                            non_null_values = row.dropna()
                            if len(non_null_values) > 0:
                                sample_values = dict(list(non_null_values.items())[:5])
                                print(f"     Row {idx}: {sample_values}")
                else:
                    print("   Sheet appears to be empty")
                    
            except Exception as e:
                print(f"   Error reading sheet data: {e}")
            
            print()
        
        wb.close()
        
    except Exception as e:
        print(f"Error analyzing file: {e}")
        return False
    
    return True

def main():
    file_path = "/Users/marcusberley/Desktop/Financial Dashboard.xlsx"
    
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        return
    
    analyze_excel_file(file_path)

if __name__ == "__main__":
    main()